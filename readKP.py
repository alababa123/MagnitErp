from collections import namedtuple
from os import PRIO_USER, curdir, name
import sys
from typing import Text
from openpyxl import load_workbook
import openpyxl
from openpyxl.cell.cell import Cell

# Типы таблиц. Большая и малая. Подробнее далее.
class TableType:
    Big, Small, UNKNOWN = range(3)

# Начало таблицы обозначается ячейкой, содержащей в себе символ №. Далее это ячейка именуется ячейкой начала таблицы.
# Правее от этого знака на той же строке должны располагаться именования столбцов (непустные ячейки).
# Далее программа построчно разбирает таблицу, опираясь на значения цвета и/или значения заполенности ячейки под ячейкой начала таблицы.



from pprintpp import pprint as pp # optional

TASK_COLOR_RGB = "FFFFFF00" # жёлтый цвет для ячеек читаемых задач 
NO_TASK_COLOR_RGB = "00000000" # белый цвет - цвет ячейки не задачи

CellCoord = namedtuple("CellCoord", ["raw", "col"])

def getCellValue(ws, coord):
    return ws.cell(coord.raw, coord.col).value

def getCellColor(ws, coord):
    return ws.cell(coord.raw, coord.col).fill.color.rgb

def printError(msg, coord):
    pass # TODO


def open_first_sheet(filename):
    wb = load_workbook(filename = filename, data_only=True)
    sheets = wb.sheetnames
    if len(sheets) > 1:
        print(f"В переданном файле обнаружено несколько страниц.\
        Будет обработана только первая из них :{sheets[0]}.")
    return wb[sheets[0]]


def check_table_begin(ws, raw):
    if ws.cell(raw, 1).value is None:
        return False
    if "№" in str(ws.cell(raw, 1).value):
        return True
    return False

# Читаем названия столбцов в таблице, пока не наткнемся на пустую ячейку
def read_table_naming(ws, begin_coord):
    names = []
    offset = 1
    while True:
        data_in_cell = ws.cell(begin_coord.raw, begin_coord.col+offset).value
        if data_in_cell is None:
            break
        names.append(data_in_cell)
        offset = offset+1
    return names

# Читаем заголовок задачи
def read_task_name(ws, coord):
    # Проверяем цвет ячейки на желтизну
    if ws.cell(coord.raw, coord.col+1).fill.fgColor.rgb != TASK_COLOR_RGB:
        return None
    if ws.cell(coord.raw, coord.col).value is not None:
        print("Ошибка. В заголовке наименований работ есть первая ячейка в строке не пуста.", coord)
        return None
    if ws.cell(coord.raw, coord.col+1).value is None:
        print("Ошибка. Отсутствует название наименований работ во второй ячейке строки.", coord)
        return None
    task_name = ws.cell(coord.raw, coord.col+1).value
    return task_name

def parseTableType(ws, table_begin_coord):
    if ws.cell(table_begin_coord.raw+1, table_begin_coord.col).fill.fgColor.rgb != NO_TASK_COLOR_RGB:
        return TableType.Big
    elif ws.cell(table_begin_coord.raw-1, table_begin_coord.col).fill.fgColor.rgb == TASK_COLOR_RGB:
        return TableType.Small
    else:
        return TableType.UNKNOWN
    

def read_table_tasks(ws, begin_coord):
    naming = read_table_naming(ws, begin_coord)
    if len(naming) == 0:
        print("Ошибка при чтении значений столбцов.", begin_coord)
        quit()
    table_tasks = []
    last_raw = 0
    typeTabe = parseTableType(ws, begin_coord)

    if typeTabe == TableType.Big:
        cur_coord = CellCoord(begin_coord.raw+1, begin_coord.col)

        while True:
            task_name = read_task_name(ws, cur_coord)
            if task_name is None:
                break
            readed_tasks, new_last_raw = read_sub_tasks(ws, CellCoord(cur_coord.raw+1, cur_coord.col), task_name, naming)
            last_raw = new_last_raw
            table_tasks.append(readed_tasks)
            cur_coord = CellCoord(last_raw, cur_coord.col)

    elif typeTabe == TableType.Small:
        # Читаем надйенную выше жёлтую строку и название задачи
        task_name = read_task_name(ws, CellCoord(begin_coord.raw-1, begin_coord.col))
        readed_task, new_last_raw = read_sub_tasks(ws, CellCoord(begin_coord.raw+1, begin_coord.col), task_name, naming)
        table_tasks.append(readed_task)
        last_raw = new_last_raw
    else:
        print("Найденная таблица не соотвутсвует критериями", begin_coord)
        quit(1)

    return table_tasks, last_raw

# Читаем строки с заданиями. Первая ячейка в строке не пустая - там лежит п/п.
def read_sub_tasks(ws, coord, name, naming):
    # Далее читаем тело задачи. Читаем данные построчно, пока не наткнемся на пустой номер п/п.
    cur_raw = coord.raw
    tasks = []
    skipEmptyRaw = True # скипием одиночные пустые строки


    while True:
        # print("While in read sub tasks:", name, cur_raw, type(ws.cell(cur_raw, coord.col).fill.fgColor.rgb))
        if type(ws.cell(cur_raw, coord.col).fill.fgColor.rgb) != str: # КОСТЫЛЬ. баг в библиотееке? при доступе к ячейке цвет не проинициализирован
            print("Некорректное значение цвета ячейки. Если здесь должен быть белый цвет, поставьте цвет заливки 'Нет цвета'.")
        if ws.cell(cur_raw, coord.col).fill.fgColor.rgb != NO_TASK_COLOR_RGB:
            break
        npp = ws.cell(cur_raw, coord.col).value
        if npp is None:
            if skipEmptyRaw:
                skipEmptyRaw = False
                cur_raw += 1
                continue
            else:
                break
        else:
            skipEmptyRaw = True

        if str(npp).isnumeric() == False:
            print(npp)
        task_data = []
        for i in range(len(naming)):
            new_column_data = ws.cell(cur_raw, coord.col+1+i).value
            task_data.append(new_column_data)
        tasks.append(task_data)
        cur_raw += 1

    # Формируем абэкт 
    task = {"name": name, "naming": naming, "inner_tasks": tasks}
    
    return task, cur_raw


def get_data_excel(file_excel_name):
    ws = open_first_sheet(file_excel_name)
    all_tasks = None
    cur_raw = 1
    while cur_raw <= ws.max_row:
        if check_table_begin(ws, cur_raw):
            table_tasks, last_raw = read_table_tasks(ws, CellCoord(cur_raw, 1))
            cur_raw = last_raw
            if all_tasks is None:
                all_tasks = table_tasks
            else:
                for new_task in table_tasks:
                    all_tasks.append(new_task)
        else:
            cur_raw += 1

    return all_tasks



if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Ождается название файла в качестве аргумента.")
        quit(1)

    try:
        # ws - work sheet - текущая страница документа, откуда читаем данные
        ws = open_first_sheet(sys.argv[1])
    except:
        print("Ошибка при открытии файла. Проверьте название и путь.")
        quit(1)

    all_tasks = None
    cur_raw = 1
    while cur_raw <= ws.max_row:
        if check_table_begin(ws, cur_raw):
            table_tasks, last_raw = read_table_tasks(ws, CellCoord(cur_raw, 1))
            cur_raw = last_raw
            if all_tasks is None:
                all_tasks = table_tasks
            else:
                for new_task in table_tasks:
                    all_tasks.append(new_task)
        else:
            cur_raw += 1

    # print(ws.max_row)

    # pp(all_tasks)
    tasks_num = 0
    for task in all_tasks:
        # print(task["name"])
        tasks_num += 1
        for inner_task in task["inner_tasks"]:
            # print(inner_task[0])
            tasks_num += 1
    print("Tasks read:", tasks_num)

    # print("bg:", ws.cell(2,1).fill.bgColor.rgb)
    # print("fg:", ws.cell(2,1).fill.fgColor.rgb)
    # print("font:", ws.cell(2,1).font.color.rgb)
    # print("value:", ws.cell(2,1).value)
    # if type(ws.cell(2,1).fill.fgColor.rgb) != str:
    #     print("style error")
